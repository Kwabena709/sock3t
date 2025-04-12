from customtkinter import *
from tkinter import messagebox
from tkinter import ttk  # For Treeview widget
from tkinter import Scrollbar  # For adding scrollbars
import mysql.connector
from PIL import Image
from customtkinter import CTkImage
import openpyxl
from tkinter import filedialog


def Export_data():
    # Get all data from the Treeview
    rows = tree.get_children()
    if not rows:
        messagebox.showinfo("No Data", "There is no data to export.")
        return

    # Ask user where to save the Excel file
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             title="Save as")

    if not file_path:
        return  # User cancelled save dialog

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Welfare Members"

        # Add headers
        headers = ["ID", "Name", "Year", "Month", "Amount"]
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_index, value=header)

        # Add row data
        for row_index, row_id in enumerate(rows, start=2):
            values = tree.item(row_id)["values"]
            for col_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

        workbook.save(file_path)
        messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to export data: {e}")

# Function to establish connection to MySQL (for creating database)
def get_initial_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="1989"
    )

# Function to connect to the welfare database (after creation)
def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="1989",
        database="welfare"
    )

# Function to create database and table if they don't exist
def setup_database():
    try:
        conn = get_initial_connection()
        cursor = conn.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS welfare")
        cursor.close()
        conn.close()

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS members (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                year INT NOT NULL,
                month VARCHAR(20) NOT NULL,
                amount DECIMAL(10, 2) NOT NULL
            )
        """)
        cursor.close()
        conn.close()
    except Exception as e:
        messagebox.showerror("Database Error", f"Failed to set up database: {e}")

# Setup database before launching the UI
setup_database()

# ------------------------ UI SETUP -------------------------

root = CTk()
root.title('Welfare Page')
root.geometry('800x650+290+25')
root.resizable(False, False)
root.configure(fg='#410C5A')
# Load the image using CTkImage
church_logo_image = CTkImage(Image.open("church logo.png"), size=(100, 100))  # Adjust size as needed

# Set custom icon (ensure .ico file exists in the same directory)
try:
    root.iconbitmap("church_logo.ico")
except Exception as e:
    print(f"Icon not set: {e}")  # Or use pass if you don't want any output


root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)


# Function to clear the Treeview
def clear_treeview():
    for row in tree.get_children():
        tree.delete(row)

# Function to refresh the Treeview
def refresh_treeview(data):
    clear_treeview()
    for row in data:
        tree.insert("", "end", values=row)

def Add_Member():
    name = NameEntry.get()
    year = YearCombobox.get()
    month = MonthCombobox.get()
    amount = AmountEntry.get()

    if not name or not year or not month or not amount:
        messagebox.showerror("Error", "Please fill in all fields.")
        return

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO members (name, year, month, amount) VALUES (%s, %s, %s, %s)",
                       (name, int(year), month, float(amount)))
        conn.commit()
        cursor.close()
        conn.close()

        messagebox.showinfo("Success", "Member added successfully.")
        Show_Members()

        NameEntry.delete(0, "end")
        YearCombobox.set("Select a Year")
        MonthCombobox.set("Select a Month")
        AmountEntry.delete(0, "end")

    except Exception as e:
        messagebox.showerror("Error", f"Error adding member: {e}")

def Delete_Member():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showerror("Error", "Please select a member to delete.")
        return

    member_id = tree.item(selected_item)["values"][0]

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM members WHERE id = %s", (member_id,))
        conn.commit()
        cursor.close()
        conn.close()

        messagebox.showinfo("Success", "Member deleted successfully.")
        Show_Members()

        NameEntry.delete(0, "end")
        YearCombobox.set("Select a Year")
        MonthCombobox.set("Select a Month")
        AmountEntry.delete(0, "end")

    except Exception as e:
        messagebox.showerror("Error", f"Error deleting member: {e}")

def Update_Member():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showerror("Error", "Please select a member to update.")
        return

    member_id = tree.item(selected_item)["values"][0]
    name = NameEntry.get()
    year = YearCombobox.get()
    month = MonthCombobox.get()
    amount = AmountEntry.get()

    if not name or not year or not month or not amount:
        messagebox.showerror("Error", "Please fill in all fields.")
        return

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(""" 
            UPDATE members
            SET name = %s, year = %s, month = %s, amount = %s
            WHERE id = %s
        """, (name, int(year), month, float(amount), member_id))
        conn.commit()
        cursor.close()
        conn.close()

        messagebox.showinfo("Success", "Member updated successfully.")
        Show_Members()

        NameEntry.delete(0, "end")
        YearCombobox.set("Select a Year")
        MonthCombobox.set("Select a Month")
        AmountEntry.delete(0, "end")

    except Exception as e:
        messagebox.showerror("Error", f"Error updating member: {e}")

def Show_Members():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM members")
        members = cursor.fetchall()
        cursor.close()
        conn.close()

        refresh_treeview(members)
    except Exception as e:
        messagebox.showerror("Error", f"Error fetching members: {e}")

def Search_Member():
    search_term = SearchEntry.get()
    if not search_term:
        messagebox.showerror("Error", "Please enter a name or year to search.")
        return

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM members
            WHERE name LIKE %s OR year LIKE %s
        """, (f"%{search_term}%", f"%{search_term}%"))
        search_results = cursor.fetchall()
        cursor.close()
        conn.close()

        SearchEntry.delete(0, "end")

        if search_results:
            refresh_treeview(search_results)
            message = f"Found {len(search_results)} member{'s' if len(search_results) > 1 else ''}."
            messagebox.showinfo("Search Results", message)
        else:
            messagebox.showinfo("Search Results", "No members found with that name or year.")

    except Exception as e:
        messagebox.showerror("Error", f"Error searching for members: {e}")

def on_treeview_select(event):
    selected_item = tree.selection()
    if selected_item:
        values = tree.item(selected_item)["values"]
        NameEntry.delete(0, "end")
        NameEntry.insert(0, values[1])
        YearCombobox.set(str(values[2]))
        MonthCombobox.set(values[3])
        AmountEntry.delete(0, "end")
        AmountEntry.insert(0, str(values[4]))

# UI elements setup
frame = CTkFrame(root)
frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

frame.grid_rowconfigure(0, weight=1)
frame.grid_columnconfigure(0, weight=1)

Namelabel = CTkLabel(frame, text='Name:', font=('bookman old style', 20))
Namelabel.grid(row=1, column=0, padx=30, pady=10, sticky='w')

NameEntry = CTkEntry(frame, placeholder_text='Enter Member Name:', font=('bookman old style', 15), width=250)
NameEntry.grid(row=1, columnspan=3, pady=10)

Yearlabel = CTkLabel(frame, text='Year:', font=('bookman old style', 20))
Yearlabel.grid(row=2, column=0, padx=30, pady=10, sticky='w')

years = [str(year) for year in range(2016, 2046)]
YearCombobox = CTkComboBox(frame, values=years, font=('bookman old style', 15), width=250, state='readonly')
YearCombobox.set("Select a Year")
YearCombobox.grid(row=2, columnspan=3, pady=10)

Monthlabel = CTkLabel(frame, text='Month:', font=('bookman old style', 20))
Monthlabel.grid(row=3, column=0, padx=30, pady=10, sticky='w')

months = ["January", "February", "March", "April", "May", "June", "July", "August",
          "September", "October", "November", "December"]
MonthCombobox = CTkComboBox(frame, values=months, font=('bookman old style', 15), width=250, state='readonly')
MonthCombobox.set("Select a Month")
MonthCombobox.grid(row=3, columnspan=3, pady=10)

Amountlabel = CTkLabel(frame, text='Amount:', font=('bookman old style', 20))
Amountlabel.grid(row=4, column=0, padx=30, pady=10, sticky='w')

AmountEntry = CTkEntry(frame, placeholder_text='Enter Amount:\u20B5', font=('bookman old style', 15), width=250)
AmountEntry.grid(row=4, columnspan=3, pady=10)

SearchLabel = CTkLabel(frame, text='Search:', font=('bookman old style', 20))
SearchLabel.grid(row=5, column=0, padx=30, pady=10, sticky='w')

SearchEntry = CTkEntry(frame, placeholder_text='Search by Name or Year', font=('bookman old style', 15), width=250)
SearchEntry.grid(row=5, columnspan=3, pady=10)

Add_member_button = CTkButton(frame, text="Add Member", font=('bookman old style', 15),
                              hover_color='red4', fg_color='brown4', command=Add_Member)
Add_member_button.grid(row=3, column=2, pady=10, padx=10, sticky="ew")

Export_data_Button = CTkButton(frame, text="Export Data", font=('bookman old style', 15),
                              hover_color='red4', fg_color='brown4', command=Export_data)
Export_data_Button.grid(row=3, column=3, pady=10, padx=10, sticky="ew")

Delete_member_button = CTkButton(frame, text="Delete Member", font=('bookman old style', 15),
                                 hover_color='red4', fg_color='brown4', command=Delete_Member)
Delete_member_button.grid(row=4, column=3, pady=10, padx=10, sticky="ew")

Update_member_button = CTkButton(frame, text="Update Member", font=('bookman old style', 15),
                                 hover_color='red4', fg_color='brown4', command=Update_Member)
Update_member_button.grid(row=4, column=2, pady=10, padx=10, sticky="ew")

Show_member_button = CTkButton(frame, text="Show Members", font=('bookman old style', 15),
                               hover_color='red4', fg_color='brown4', command=Show_Members)
Show_member_button.grid(row=5, column=3, pady=10, padx=10, sticky="ew")

Search_member_button = CTkButton(frame, text="Search Member", font=('bookman old style', 15),
                                 hover_color='red4', fg_color='brown4', command=Search_Member)
Search_member_button.grid(row=5, column=2, pady=10, padx=10, sticky="ew")


label_name=CTkLabel(frame,text='Vbci Abundant Faith_Welfare', font=('bookman old style', 30,'bold'), text_color='brown4')
label_name.place(x=25,y=10)

logo_label = CTkLabel(frame, image=church_logo_image, text="")  # text="" to hide default label text
logo_label.place(x=450, y=80)  # Adjust x and y to position the logo nicely


# Treeview setup
treeview_frame = CTkFrame(frame)
treeview_frame.grid(row=7, column=0, columnspan=5, padx=20, pady=20, sticky="nsew")
treeview_frame.grid_rowconfigure(0, weight=1)
treeview_frame.grid_columnconfigure(0, weight=1)

columns = ("ID", "Name", "Year", "Month", "Amount")
tree = ttk.Treeview(treeview_frame, columns=columns, show="headings", height=10)
tree.heading("ID", text="ID")
tree.heading("Name", text="Name")
tree.heading("Year", text="Year")
tree.heading("Month", text="Month")
tree.heading("Amount", text="Amount \u20B5")

tree.column("ID", width=50, anchor="center")
tree.column("Name", width=150, anchor="w")
tree.column("Year", width=100, anchor="center")
tree.column("Month", width=150, anchor="w")
tree.column("Amount", width=100, anchor="center")

v_scrollbar = Scrollbar(treeview_frame, orient="vertical", command=tree.yview)
v_scrollbar.grid(row=0, column=1, sticky="ns", padx=(5, 0))

h_scrollbar = Scrollbar(treeview_frame, orient="horizontal", command=tree.xview)
h_scrollbar.grid(row=1, column=0, sticky="ew", padx=(5, 0))

tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
tree.grid(row=0, column=0, sticky="nsew")

tree.bind("<<TreeviewSelect>>", on_treeview_select)

root.mainloop()
