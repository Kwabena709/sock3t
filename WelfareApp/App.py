from customtkinter import *
from tkinter import messagebox
from tkinter import ttk
from tkinter import Scrollbar
import mysql.connector
from PIL import Image
from customtkinter import CTkImage
import openpyxl
from tkinter import filedialog


# ----------------- Theme Toggle ------------------

def apply_theme_mode():
    current_mode = get_appearance_mode()
    if current_mode.lower() == "dark":
        ThemeToggleButton.configure(
            text="Light Mode",
            fg_color="white",
            text_color="black",
            hover_color="white"  # Prevents hover color change
        )
        label_name.configure(text_color="white")
    else:
        ThemeToggleButton.configure(
            text="Dark Mode",
            fg_color="black",
            text_color="white",
            hover_color="black"  # Prevents hover color change
        )
        label_name.configure(text_color="#410C5A")

def toggle_theme():
    current_mode = get_appearance_mode()
    if current_mode.lower() == "dark":
        set_appearance_mode("light")
    else:
        set_appearance_mode("dark")
    apply_theme_mode()

# ----------------- Functions ------------------

def Export_data():
    rows = tree.get_children()
    if not rows:
        messagebox.showinfo("No Data", "There is no data to export.")
        return

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             title="Save as")
    if not file_path:
        return

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Welfare Members"

        headers = ["ID", "Name", "Year", "Month", "Amount"]
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_index, value=header)

        for row_index, row_id in enumerate(rows, start=2):
            values = tree.item(row_id)["values"]
            for col_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

        workbook.save(file_path)
        messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to export data: {e}")

def get_initial_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="1989"
    )

def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="1989",
        database="welfare"
    )

def setup_database():
    try:
        conn = get_initial_connection()
        cursor = conn.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS welfare")
        cursor.close()
        conn.close()

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(""" 
            CREATE TABLE IF NOT EXISTS members (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                year INT NOT NULL,
                month VARCHAR(20) NOT NULL,
                amount DECIMAL(10, 2) NOT NULL
            )
        """)
        cursor.close()
        conn.close()
    except Exception as e:
        messagebox.showerror("Database Error", f"Failed to set up database: {e}")

setup_database()

# ----------------- UI Setup ------------------

root = CTk()
root.title('Welfare Page')
root.geometry('1250x650+45+25')
root.resizable(False, False)
root.configure(fg='#410C5A')

church_logo_image = CTkImage(Image.open("church logo.png"), size=(100, 100))

try:
    root.iconbitmap("church_logo.ico")
except Exception as e:
    print(f"Icon not set: {e}")

root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

# ----------------- UI Functions ------------------

def clear_treeview():
    for row in tree.get_children():
        tree.delete(row)

def refresh_treeview(data):
    clear_treeview()
    for row in data:
        tree.insert("", "end", values=row)

def Add_Member():
    name = NameEntry.get()
    year = YearCombobox.get()
    month = MonthCombobox.get()
    amount = AmountEntry.get()

    if not name or not year or not month or not amount:
        messagebox.showerror("Error", "Please fill in all fields.")
        return

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO members (name, year, month, amount) VALUES (%s, %s, %s, %s)",
                       (name, int(year), month, float(amount)))
        conn.commit()
        cursor.close()
        conn.close()

        messagebox.showinfo("Success", "Member added successfully.")
        Show_Members()

        NameEntry.delete(0, "end")
        YearCombobox.set("Select a Year")
        MonthCombobox.set("Select a Month")
        AmountEntry.delete(0, "end")

    except Exception as e:
        messagebox.showerror("Error", f"Error adding member: {e}")

def Delete_Member():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showerror("Error", "Please select a member to delete.")
        return
    if not messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this member?"):
        return

    member_id = tree.item(selected_item)["values"][0]

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM members WHERE id = %s", (member_id,))
        conn.commit()
        cursor.close()
        conn.close()

        messagebox.showinfo("Success", "Member deleted successfully.")
        Show_Members()

        NameEntry.delete(0, "end")
        YearCombobox.set("Select a Year")
        MonthCombobox.set("Select a Month")
        AmountEntry.delete(0, "end")

    except Exception as e:
        messagebox.showerror("Error", f"Error deleting member: {e}")

def Update_Member():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showerror("Error", "Please select a member to update.")
        return

    member_id = tree.item(selected_item)["values"][0]
    name = NameEntry.get()
    year = YearCombobox.get()
    month = MonthCombobox.get()
    amount = AmountEntry.get()

    if not name or not year or not month or not amount:
        messagebox.showerror("Error", "Please fill in all fields.")
        return

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(""" 
            UPDATE members
            SET name = %s, year = %s, month = %s, amount = %s
            WHERE id = %s
        """, (name, int(year), month, float(amount), member_id))
        conn.commit()
        cursor.close()
        conn.close()

        messagebox.showinfo("Success", "Member updated successfully.")
        Show_Members()

        NameEntry.delete(0, "end")
        YearCombobox.set("Select a Year")
        MonthCombobox.set("Select a Month")
        AmountEntry.delete(0, "end")

    except Exception as e:
        messagebox.showerror("Error", f"Error updating member: {e}")

def Show_Members():
    root.update()  # Ensure the loading indicator appears
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM members")
        members = cursor.fetchall()
        cursor.close()
        conn.close()

        if members:
            refresh_treeview(members)
            messagebox.showinfo("Members Found", f"Found {len(members)} member{'s' if len(members) > 1 else ''}.")
        else:
            messagebox.showinfo("No Members", "No members found in the database.")

    except Exception as e:
        messagebox.showerror("Error", f"Error fetching members: {e}")


def Search_Member():
    search_term = SearchEntry.get()
    if not search_term:
        messagebox.showerror("Error", "Please enter a name or year to search.")
        return

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(""" 
            SELECT * FROM members
            WHERE name LIKE %s OR year LIKE %s
        """, (f"%{search_term}%", f"%{search_term}%"))
        search_results = cursor.fetchall()
        cursor.close()
        conn.close()

        SearchEntry.delete(0, "end")

        if search_results:
            refresh_treeview(search_results)
            message = f"Found {len(search_results)} member{'s' if len(search_results) > 1 else ''}."
            messagebox.showinfo("Search Results", message)
        else:
            messagebox.showinfo("Search Results", "No members found with that name or year.")

    except Exception as e:
        messagebox.showerror("Error", f"Error searching for members: {e}")

def on_treeview_select(event):
    selected_item = tree.selection()
    if selected_item:
        values = tree.item(selected_item)["values"]
        NameEntry.delete(0, "end")
        NameEntry.insert(0, values[1])
        YearCombobox.set(str(values[2]))
        MonthCombobox.set(values[3])
        AmountEntry.delete(0, "end")
        AmountEntry.insert(0, str(values[4]))

# ----------------- Frames and Widgets ------------------

frame = CTkFrame(root)
frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

frame.grid_rowconfigure(7, weight=1)

# Entries and Labels
NameLabel = CTkLabel(frame, text='Name:', font=('bookman old style', 20))
NameLabel.grid(row=1, column=0, padx=30, pady=10, sticky='w')

NameEntry = CTkEntry(frame, placeholder_text='Enter Member Name:', font=('bookman old style', 20), height=40, width=250)
NameEntry.grid(row=1, column=1, pady=10)

YearLabel = CTkLabel(frame, text='Year:', font=('bookman old style', 20))
YearLabel.grid(row=2, column=0, padx=30, pady=10, sticky='w')

years = [str(year) for year in range(2016, 2046)]
YearCombobox = CTkComboBox(frame, values=years, font=('bookman old style', 20), height=40, width=250, state='readonly')
YearCombobox.set("Select a Year")
YearCombobox.grid(row=2, column=1, pady=10)

MonthLabel = CTkLabel(frame, text='Month:', font=('bookman old style', 20))
MonthLabel.grid(row=3, column=0, padx=30, pady=10, sticky='w')

months = ["January", "February", "March", "April", "May", "June", "July", "August",
          "September", "October", "November", "December"]
MonthCombobox = CTkComboBox(frame, values=months, font=('bookman old style', 20), height=40, width=250, state='readonly')
MonthCombobox.set("Select a Month")
MonthCombobox.grid(row=3, column=1, pady=10)

AmountLabel = CTkLabel(frame, text='Amount:', font=('bookman old style', 20))
AmountLabel.grid(row=4, column=0, padx=30, pady=10, sticky='w')

AmountEntry = CTkEntry(frame, placeholder_text='Enter Amount:GH\u20B5', font=('bookman old style', 20), height=40, width=250)
AmountEntry.grid(row=4, column=1, pady=10)

SearchLabel = CTkLabel(frame, text='Search:', font=('bookman old style', 20))
SearchLabel.grid(row=5, column=0, padx=30, pady=10, sticky='w')

SearchEntry = CTkEntry(frame, placeholder_text='Search by Name or Year', font=('bookman old style', 20), height=40, width=250)
SearchEntry.grid(row=5, column=1, pady=10)

# Buttons
Add_member_button = CTkButton(frame, text="Add Member", font=('bookman old style',20),
                              hover_color='black', fg_color='#410C5A', height=40, command=Add_Member)
Add_member_button.grid(row=3, column=2, pady=10, padx=10, sticky="ew")

Export_data_Button = CTkButton(frame, text="Export Data", font=('bookman old style', 20),
                              hover_color='black', fg_color='#410C5A', height=40, command=Export_data)
Export_data_Button.grid(row=3, column=3, pady=10, padx=10, sticky="ew")

Delete_member_button = CTkButton(frame, text="Delete Member", font=('bookman old style', 20),
                                 hover_color='black', fg_color='#410C5A', height=40, command=Delete_Member)
Delete_member_button.grid(row=4, column=3, pady=10, padx=10, sticky="ew")

Update_member_button = CTkButton(frame, text="Update Member", font=('bookman old style', 20),
                                 hover_color='black', fg_color='#410C5A', height=40, command=Update_Member)
Update_member_button.grid(row=4, column=2, pady=10, padx=10, sticky="ew")

Show_member_button = CTkButton(frame, text="Show Members", font=('bookman old style', 20),
                               hover_color='black', fg_color='#410C5A', height=40, command=Show_Members)
Show_member_button.grid(row=5, column=3, pady=10, padx=10, sticky="ew")

Search_member_button = CTkButton(frame, text="Search Member", font=('bookman old style', 20),
                                 hover_color='black', fg_color='#410C5A', height=40, command=Search_Member)
Search_member_button.grid(row=5, column=2, pady=10, padx=10, sticky="ew")


# Theme Toggle Button
ThemeToggleButton = CTkButton(root, text="", font=('bookman old style', 15),
                              command=toggle_theme, width=120)
ThemeToggleButton.place(x=1085, y=30)

# Logo
logo_label = CTkLabel(frame, image=church_logo_image, text="")
logo_label.place(x=530, y=20)


label_name = CTkLabel(root, text='------Vbci Abundant Faith Sanctuary-Welfare------', font=('Times New Roman', 30,'bold'))
label_name.place(x=300,y=350)

# Treeview
treeview_frame = CTkFrame(root)
treeview_frame.grid(row=1, column=0, columnspan=5, padx=20, pady=20, sticky="nsew")
treeview_frame.grid_rowconfigure(0, weight=1)
treeview_frame.grid_columnconfigure(0, weight=1)

columns = ("ID", "Name", "Year", "Month", "Amount")
tree = ttk.Treeview(treeview_frame, columns=columns, show="headings", height=10)

# Custom style for treeview with light grey background
style = ttk.Style()
style.theme_use("default")  # Allow custom styling

style.configure("Custom.Treeview",
                background="#D3D3D3",         # Light grey background
                fieldbackground="#D3D3D3",    # Background of the table area
                foreground="black",
                font=('Times New Roman', 14))

style.configure("Custom.Treeview.Heading",
                font=('Times New Roman', 16, 'bold'),
                foreground='red',
                background='#D3D3D3')         # Optional: light grey heading

style.layout("Custom.Treeview.Heading", [
    ('Treeheading.cell', {'sticky': 'nswe', 'border': '2'}),
    ('Treeheading.border', {'sticky': 'nswe'}),
    ('Treeheading.padding', {'sticky': 'nswe'}),
    ('Treeheading.image', {'side': 'right', 'sticky': ''}),
    ('Treeheading.text', {'sticky': 'we'})
])

style.configure("Custom.Treeview.Heading",
                font=('Times New Roman', 16, 'bold'),
                foreground='red',
                background='#D3D3D3',
                relief="raised",       # Adds border effect
                borderwidth=2)         # Thickness of the border

tree.configure(style="Custom.Treeview")

# Set column headings
for col in columns:
    tree.heading(col, text=col)
tree.column("ID", width=50, anchor="center")
tree.column("Name", width=150, anchor="w")
tree.column("Year", width=100, anchor="center")
tree.column("Month", width=150, anchor="w")
tree.column("Amount", width=100, anchor="center")

# Scrollbars
v_scrollbar = Scrollbar(treeview_frame, orient="vertical", command=tree.yview)
v_scrollbar.grid(row=0, column=1, sticky="ns", padx=(5, 0))

h_scrollbar = Scrollbar(treeview_frame, orient="horizontal", command=tree.xview)
h_scrollbar.grid(row=1, column=0, sticky="ew", pady=(5, 0))
tree.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
tree.grid(row=0, column=0, sticky="nsew")

tree.bind("<<TreeviewSelect>>", on_treeview_select)


# ----------------- Final UI Updates ------------------

apply_theme_mode()


root.mainloop()
